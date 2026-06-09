import random
import csv
import openpyxl
from typing import List, Dict, Tuple, Any, TextIO

def read_meds(filename: str) -> List[str]:
    """
    Reads a text file and extracts the drug names into a list.

    Parameters
    ---------
    filename : str
        The path to the text file to be read in the file system.

    Returns
    -------
    list of str
        A list containing the clean drug names without extra spaces.

    Raises
    ------
    FileNotFoundError
        If the specified text file is not found.

    OSError
        If any other unexpected error occurs while opening or reading the file.
    """
    clean_list: List[str] = []
    try:
        file_handler: TextIO
        with open(filename, 'r', encoding='utf-8') as file_handler:
            line: str
            for line in file_handler:
                clean_text: str = line.strip()
                if clean_text:
                    clean_list.append(clean_text)
        return clean_list
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Erro: O ficheiro '{filename}' não foi encontrado.") from exc
    except Exception as exc:
        raise OSError(f"Erro inesperado ao ler o ficheiro: {exc}") from exc

def gen_matrix(meds: List[str]) -> Dict[Tuple[str, str], int]:
    """
    Gera uma matriz de interações medicamentosas estruturada em dicionário e tuplos.

    Parameters
    ----------
    meds : list of str
        A lista com os nomes de todos os medicamentos extraídos.

    Returns
    -------
    dict
        Dicionário cujas chaves são tuplos de pares de medicamentos (estrutura imutável) 
        e os valores são o grau de interação (0 para o mesmo medicamento, 1 a 6 para distintos).
    """
    matrix: Dict[Tuple[str, str], int] = {}
    med_row: str
    med_col: str
    for med_row in meds:
        for med_col in meds:
            if med_row == med_col:
                matrix[(med_row, med_col)] = 0
            else:
                matrix[(med_row, med_col)] = random.randint(1, 6)
    return matrix

def export_excel(matrix: Dict[Tuple[str, str], int], meds: List[str], filename: str) -> None:
    """
    Exports the interaction matrix data to an Excel spreadsheet.

    Parameters
    ---------
    matrix : dict
        The generated interaction dictionary with combinations of all medications.
    meds : list of str
        The list of medication names to be used in the header and first column .
    filename : str
        The name or directory of the destination file where the Excel document will be saved.

    Raises
    ----
    TypeError
        If the submitted matrix is ​​null or uninitialized.

    OSError
        If there are permission problems, a corrupted file, or errors saving to disk.
    """
    if matrix is None:
        raise TypeError("Erro Crítico: A matriz de dados não pode ser nula.")
    
    try:
        workbook: openpyxl.Workbook = openpyxl.Workbook()
        worksheet: Any = workbook.active
        
        header_row: List[str] = [""]
        med: str
        for med in meds:
            header_row.append(med)
        
        worksheet.append(header_row)
        
        row_med: str
        col_med: str
        for row_med in meds:
            new_row: List[Any] = [row_med]
            for col_med in meds:
                effect: int = matrix[(row_med, col_med)]
                new_row.append(effect)
            
            worksheet.append(new_row)
            
        workbook.save(filename)
    except Exception as exc:
        raise OSError(f"Erro ao tentar criar ou guardar o ficheiro Excel: {exc}") from exc

def convert_xlsx_to_csv(excel_path: str, csv_path: str) -> None:
    """
    Converts an Excel file to a rigorously formatted CSV file.
    The function uses the openpyxl library to load the active sheet 
    of an Excel document and immediately converts it, writing the content 
    line by line to a destination file natively supported by the CSV format.

    Parameters
    ----------
    excel_path : str 
    Exact path on disk of the source Excel file.

    csv_path : str 
    Path of the CSV file to be created as the destination.

    Raises
    ------
    FileNotFoundError
    If the indicated Excel file is not found on the system.

    IOError
    If any input/output or permissions anomaly occurs.
    """
    try:
        workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_path)
        worksheet: Any = workbook.active
        
        csv_file: TextIO
        with open(csv_path, 'w', encoding='utf-8', newline='') as csv_file:
            csv_writer: Any = csv.writer(csv_file, delimiter=',')
            
            row: Tuple[Any, ...]
            for row in worksheet.iter_rows(values_only=True):
                row_data: List[Any] = list(row)
                csv_writer.writerow(row_data)

    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Erro: O ficheiro Excel de origem '{excel_path}' não foi encontrado.") from exc
  